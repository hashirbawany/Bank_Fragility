# Section 1: Package loading, directories, and report date
import os
import sys
import zipfile
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy.stats.mstats import winsorize

load_dotenv(Path(__file__).resolve().parent.parent / ".env")

sys.path.insert(0, str(Path(__file__).resolve().parent))
from pull_gsib_banks import pull_gsib_list

DATA_DIR   = Path(__file__).resolve().parent.parent / "_data"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "_output"
REPORT_DATE_SLASH = os.getenv("REPORT_DATE_SLASH", "12/31/2025")  # MM/DD/YYYY — e.g. 03/31/2025
REPORT_DATE       = REPORT_DATE_SLASH.replace("/", "")

DATA_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

ZIP_PATH = DATA_DIR / f"FFIEC CDR Call Bulk All Schedules {REPORT_DATE}.zip"
if not ZIP_PATH.exists():
    raise FileNotFoundError(f"Missing FFIEC zip file: {ZIP_PATH}")


# Section 2: Defining helper functions
def read_ffiec(zf, filename):
    with zf.open(filename) as f:
        df = pd.read_csv(
            f,
            sep="\t",
            header=0,
            skiprows=[1],
            low_memory=False,
            dtype=str,
        )
    df.columns = df.columns.str.strip().str.replace('"', "").str.lower()
    df = df.rename(columns={"idrssd": "rssd9001"})
    df["rssd9001"] = pd.to_numeric(df["rssd9001"], errors="coerce")
    df = df.dropna(subset=["rssd9001"])
    df["rssd9001"] = df["rssd9001"].astype(int)
    df = df.set_index("rssd9001")
    df = df.apply(pd.to_numeric, errors="coerce")
    return df


def fmt_dollar(num_thousands):
    """Format a number in thousands to B/T string."""
    num = num_thousands * 1000
    if num >= 1e12:
        return f"{num / 1e12:.1f}T"
    elif num >= 1e9:
        return f"{num / 1e9:.1f}B"
    else:
        return f"{num / 1e6:.1f}M"


def find_member_name(zf, target_stub: str) -> str:
    names = zf.namelist()
    matches = [name for name in names if target_stub.lower() in name.lower()]
    if not matches:
        raise FileNotFoundError(
            f"Could not find '{target_stub}' inside zip.\nAvailable files:\n"
            + "\n".join(names[:50])
        )
    return matches[0]


def check_cols(df, cols, df_name):
    missing = [c for c in cols if c not in df.columns]
    if missing:
        print(f"  MISSING {df_name}: {missing}")
    else:
        print(f"  OK {df_name} - all present")


def winsorized_mean_sd(
    df: pd.DataFrame, asset_col: str = "Total Asset"
) -> tuple[pd.Series, pd.Series]:
    """
    Compute winsorized mean and sd of each liability category as a percent of total assets.
    Returns (mean_series, sd_series).
    """
    if df.empty:
        cols = [c for c in df.columns if c not in ["Bank Category", asset_col]]
        empty = pd.Series(index=cols, dtype=float)
        return empty, empty

    value_cols = [c for c in df.columns if c not in ["Bank Category", asset_col]]
    ratio_df = df[value_cols].div(df[asset_col], axis=0) * 100

    mean_series = ratio_df.apply(
        lambda x: pd.Series(
            np.array(winsorize(x.dropna(), limits=[0.05, 0.05])), dtype=float
        ).mean()
    )

    sd_series = ratio_df.apply(
        lambda x: pd.Series(
            np.array(winsorize(x.dropna(), limits=[0.05, 0.05])), dtype=float
        ).std()
    )

    return mean_series, sd_series


def large_num(num):
    num = num * 1000
    if num < 1_000_000_000:
        return f"{num / 1_000_000_000:.1f}B" if num >= 100_000_000 else str(num)
    else:
        return (
            f"{num / 1_000_000_000_000:.1f}T"
            if num >= 1_000_000_000_000
            else f"{num / 1_000_000_000:.1f}B"
        )


def allocate_across_buckets(df, source_col, prefix, weights):
    total_weight = sum(weights.values())
    if not np.isclose(total_weight, 1.0):
        raise ValueError(f"Weights for {prefix} must sum to 1. Got {total_weight}")

    source = df[source_col].fillna(0)
    for bucket, weight in weights.items():
        df[f"{prefix}_{bucket}"] += source * weight


def write_summary_sheet(ws, df, panel_title, start_row=1):
    """Write a summary stats DataFrame to a worksheet with formatting."""
    header_fill = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    subgroup_fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
    col_labels = [""] + list(df.columns)
    n_cols = len(col_labels)

    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=n_cols,
    )
    title_cell = ws.cell(row=start_row, column=1, value=panel_title)
    title_cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    start_row += 1

    group_headers = [
        ("", 1),
        ("Aggregate", 1),
        ("Full Sample", 2),
        ("Small Banks", 2),
        ("Large Banks", 2),
        ("GSIB", 2),
    ]
    col = 1
    for label, span in group_headers:
        if span > 1:
            ws.merge_cells(
                start_row=start_row,
                start_column=col,
                end_row=start_row,
                end_column=col + span - 1,
            )
        cell = ws.cell(row=start_row, column=col, value=label)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col += span
    start_row += 1

    detail_headers = [
        "Variable",
        "Aggregate",
        "Mean",
        "Std Dev",
        "Mean",
        "Std Dev",
        "Mean",
        "Std Dev",
        "Mean",
        "Std Dev",
    ]
    for c, label in enumerate(detail_headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=label)
        cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium"))
    start_row += 1

    shaded_rows = {
        "Total Asset $",
        "N Banks",
        "Cash",
        "Securities",
        "Total Loan",
        "Fed Funds Sold",
        "Reverse Repo",
        "Total Liability",
        "Domestic Deposit",
        "Total Equity",
    }

    for row_label, row_data in df.iterrows():
        is_shaded = row_label in shaded_rows
        fill = subgroup_fill if is_shaded else PatternFill()

        label_cell = ws.cell(row=start_row, column=1, value=row_label)
        label_cell.font = Font(name="Arial", bold=is_shaded, size=9)
        label_cell.fill = fill
        label_cell.alignment = Alignment(
            horizontal="left", vertical="center", indent=0 if is_shaded else 1
        )

        for c, val in enumerate(row_data, start=2):
            cell = ws.cell(row=start_row, column=c, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="right", vertical="center")

        start_row += 1

    return start_row


# Section 3: Loading the zip files & creating rcfd, rcon, and rcfn files
with zipfile.ZipFile(ZIP_PATH) as zf:
    rc_name = find_member_name(zf, f"FFIEC CDR Call Schedule RC {REPORT_DATE}")
    rca_name = find_member_name(zf, f"FFIEC CDR Call Schedule RCA {REPORT_DATE}")
    rcci_name = find_member_name(zf, f"FFIEC CDR Call Schedule RCCI {REPORT_DATE}")
    rce_name = find_member_name(zf, f"FFIEC CDR Call Schedule RCE {REPORT_DATE}")

    rc = read_ffiec(zf, rc_name)
    rca = read_ffiec(zf, rca_name)
    rcci = read_ffiec(zf, rcci_name)
    rce = read_ffiec(zf, rce_name)

    rcb_part_names = [
        name
        for name in zf.namelist()
        if f"FFIEC CDR Call Schedule RCB {REPORT_DATE}".lower() in name.lower()
    ]

    if len(rcb_part_names) == 2:
        rcb = pd.concat(
            [read_ffiec(zf, name) for name in sorted(rcb_part_names)], axis=1
        )
    elif len(rcb_part_names) == 1:
        rcb = read_ffiec(zf, rcb_part_names[0])
    else:
        raise FileNotFoundError(
            f"Could not find RCB files for {REPORT_DATE} in zip."
        )

rcfd_df = pd.concat(
    [
        rc[[c for c in rc.columns if c.startswith("rcfd")]],
        rca[[c for c in rca.columns if c.startswith("rcfd")]],
        rcb[[c for c in rcb.columns if c.startswith("rcfd")]],
        rcci[[c for c in rcci.columns if c.startswith("rcfd")]],
    ],
    axis=1,
)
rcfd_df = rcfd_df.loc[:, ~rcfd_df.columns.duplicated()]

rcon_df = pd.concat(
    [
        rc[[c for c in rc.columns if c.startswith("rcon")]],
        rcb[[c for c in rcb.columns if c.startswith("rcon")]],
        rcci[[c for c in rcci.columns if c.startswith("rcon")]],
        rce[[c for c in rce.columns if c.startswith("rcon")]],
    ],
    axis=1,
)
rcon_df = rcon_df.loc[:, ~rcon_df.columns.duplicated()]

rcfn_df = rc[[c for c in rc.columns if c.startswith("rcfn")]]


# Section 4: Variable groupings
global_rmbs = [
    "rcfdg301", "rcfdg303", "rcfdg305", "rcfdg307", "rcfdg309", "rcfdg311",
    "rcfdg313", "rcfdg315", "rcfdg317", "rcfdg319", "rcfdg321", "rcfdg323",
]
global_cmbs = [
    "rcfdk143", "rcfdk145", "rcfdk147", "rcfdk149", "rcfdk151", "rcfdk153", "rcfdk157",
]
global_abs = ["rcfdc988", "rcfdc027"]
global_other = ["rcfd1738", "rcfd1741", "rcfd1743", "rcfd1746"]
global_rs_loan = [
    "rcfdf158", "rcfdf159", "rcfd1420", "rcfd1420", "rcfd1797",
    "rcfd5367", "rcfd5368", "rcfd1460", "rcfdf160", "rcfdf161",
]
global_rs_residential_loan = ["rcfd1420", "rcfd1797", "rcfd5367", "rcfd5368", "rcfd1460"]
global_rs_commercial_loan = ["rcfdf160", "rcfdf161"]
global_rs_other_loan = ["rcfdf158", "rcfdf159"]
global_ci_loan = ["rcfd1763", "rcfd1764"]
global_consumer_loan = ["rcfdb538", "rcfdb539", "rcfdk137", "rcfdk207"]

domestic_cash = ["rcon0081", "rcon0071"]
domestic_total = ["rcon1771", "rcon1773"]
domestic_treasury = ["rcon0213", "rcon1287"]
domestic_rmbs = [
    "rconht55", "rconht57", "rcong309", "rcong311", "rcong313",
    "rcong315", "rcong317", "rcong319", "rcong321", "rcong323",
]
domestic_cmbs = ["rconk143", "rconk145", "rconk147", "rconk149", "rconk151", "rconk153", "rconk157"]
domestic_abs = ["rconc988", "rconc027", "rconht59", "rconht61"]
domestic_other = ["rcon1738", "rcon1741", "rcon1743", "rcon1746"]
domestic_rs_loan = [
    "rconf158", "rconf159", "rcon1420", "rcon1420", "rcon1797",
    "rcon5367", "rcon5368", "rcon1460", "rconf160", "rconf161",
]
domestic_rs_residential_loan = ["rcon1420", "rcon1797", "rcon5367", "rcon5368", "rcon1460"]
domestic_rs_commercial_loan = ["rconf160", "rconf161"]
domestic_rs_other_loan = ["rconf158", "rconf159"]
domestic_ci_loan = ["rcon1766"]
domestic_consumer_loan = ["rconb538", "rconb539", "rconk137", "rconk207"]
domestic_non_rep_loan = ["rconj454", "rconj464", "rconj451"]

insured_deposit = ["rconhk05", "rconmt91", "rconmt87"]
uninsured_long = ["rconhk14", "rconhk15"]

print("Checking rcfd_df...")
check_cols(
    rcfd_df,
    ["rcfd2170", "rcfd0010", "rcfd1771", "rcfd1773", "rcfd0213", "rcfd1287"],
    "basic cols",
)
check_cols(rcfd_df, global_rmbs, "global_rmbs")
check_cols(rcfd_df, global_cmbs, "global_cmbs")
check_cols(rcfd_df, global_abs, "global_abs")
check_cols(rcfd_df, global_other, "global_other")
check_cols(rcfd_df, global_rs_loan, "global_rs_loan")
check_cols(rcfd_df, global_rs_residential_loan, "global_rs_residential_loan")
check_cols(rcfd_df, global_rs_commercial_loan, "global_rs_commercial_loan")
check_cols(rcfd_df, global_rs_other_loan, "global_rs_other_loan")
check_cols(rcfd_df, global_ci_loan, "global_ci_loan")
check_cols(rcfd_df, global_consumer_loan, "global_consumer_loan")
check_cols(rcfd_df, ["rcfd1590", "rcfd2122", "rcfdb989"], "other rcfd")

print("\nChecking rcon_df...")
check_cols(
    rcon_df,
    ["rcon2170", "rconb987", "rconb989", "rcon2122", "rcon1590"],
    "basic cols",
)
check_cols(rcon_df, domestic_cash, "domestic_cash")
check_cols(rcon_df, domestic_total, "domestic_total")
check_cols(rcon_df, domestic_treasury, "domestic_treasury")
check_cols(rcon_df, domestic_rmbs, "domestic_rmbs")
check_cols(rcon_df, domestic_cmbs, "domestic_cmbs")
check_cols(rcon_df, domestic_abs, "domestic_abs")
check_cols(rcon_df, domestic_other, "domestic_other")
check_cols(rcon_df, domestic_rs_loan, "domestic_rs_loan")
check_cols(rcon_df, domestic_rs_residential_loan, "domestic_rs_residential_loan")
check_cols(rcon_df, domestic_rs_commercial_loan, "domestic_rs_commercial_loan")
check_cols(rcon_df, domestic_rs_other_loan, "domestic_rs_other_loan")
check_cols(rcon_df, domestic_ci_loan, "domestic_ci_loan")
check_cols(rcon_df, domestic_consumer_loan, "domestic_consumer_loan")
check_cols(rcon_df, domestic_non_rep_loan, "domestic_non_rep_loan")

print("\nChecking rcfn_df...")
check_cols(rcfn_df, ["rcfn2200"], "rcfn2200")

# Maturity bucket mappings for mark-to-market calculations
global_rmbs_buckets = {
    "lt1y": ["rcfdg301", "rcfdg303"],
    "1_3y": ["rcfdg305", "rcfdg307"],
    "3_5y": ["rcfdg309", "rcfdg311"],
    "5_10y": ["rcfdg313", "rcfdg315"],
    "10_15y": ["rcfdg317", "rcfdg319"],
    "15plus": ["rcfdg321", "rcfdg323"],
}

domestic_rmbs_buckets = {
    "lt1y": ["rconht55", "rconht57"],
    "1_3y": ["rcong309", "rcong311"],
    "3_5y": ["rcong313", "rcong315"],
    "5_10y": ["rcong317", "rcong319"],
    "10_15y": ["rcong321"],
    "15plus": ["rcong323"],
}


# Section 5: Creating Asset Table
rcfd_data = pd.DataFrame(index=rcfd_df.index)
rcfd_data["Total Asset"] = rcfd_df["rcfd2170"]
rcfd_data["cash"] = rcfd_df["rcfd0010"]
rcfd_data["security_total"] = rcfd_df["rcfd1771"] + rcfd_df["rcfd1773"]
rcfd_data["security_treasury"] = rcfd_df["rcfd0213"] + rcfd_df["rcfd1287"]
rcfd_data["security_rmbs"] = rcfd_df[global_rmbs].sum(axis=1)
rcfd_data["security_cmbs"] = rcfd_df[global_cmbs].sum(axis=1)
rcfd_data["security_abs"] = rcfd_df[global_abs].sum(axis=1)
rcfd_data["security_other"] = rcfd_df[global_other].sum(axis=1)
rcfd_data["Total_Loan"] = rcfd_df["rcfd2122"]
rcfd_data["Real_Estate_Loan"] = rcfd_df[global_rs_loan].sum(axis=1)
rcfd_data["Residential_Mortgage"] = rcfd_df[global_rs_residential_loan].sum(axis=1)
rcfd_data["Commercial_Mortgage"] = rcfd_df[global_rs_commercial_loan].sum(axis=1)
rcfd_data["Other_Real_Estate_Mortgage"] = rcfd_df[global_rs_other_loan].sum(axis=1)
rcfd_data["Agri_Loan"] = rcfd_df["rcfd1590"]
rcfd_data["Comm_Indu_Loan"] = rcfd_df[global_ci_loan].sum(axis=1)
rcfd_data["Consumer_Loan"] = rcfd_df[global_consumer_loan].sum(axis=1)
rcfd_data["Non_Rep_Loan"] = np.nan
rcfd_data["Fed_Fund_Sold"] = rcon_df["rconb987"]
rcfd_data["Reverse_Repo"] = rcfd_df["rcfdb989"]

rcon_data = pd.DataFrame(index=rcon_df.index)
rcon_data["Total Asset"] = rcon_df["rcon2170"]
rcon_data["cash"] = rcon_df[domestic_cash].sum(axis=1)
rcon_data["security_total"] = rcon_df[domestic_total].sum(axis=1)
rcon_data["security_treasury"] = rcon_df[domestic_treasury].sum(axis=1)
rcon_data["security_rmbs"] = rcon_df[domestic_rmbs].sum(axis=1)
rcon_data["security_cmbs"] = rcon_df[domestic_cmbs].sum(axis=1)
rcon_data["security_abs"] = rcon_df[domestic_abs].sum(axis=1)
rcon_data["security_other"] = rcon_df[domestic_other].sum(axis=1)
rcon_data["Total_Loan"] = rcon_df["rcon2122"]
rcon_data["Real_Estate_Loan"] = rcon_df[domestic_rs_loan].sum(axis=1)
rcon_data["Residential_Mortgage"] = rcon_df[domestic_rs_residential_loan].sum(axis=1)
rcon_data["Commercial_Mortgage"] = rcon_df[domestic_rs_commercial_loan].sum(axis=1)
rcon_data["Other_Real_Estate_Mortgage"] = rcon_df[domestic_rs_other_loan].sum(axis=1)
rcon_data["Agri_Loan"] = rcon_df["rcon1590"]
rcon_data["Comm_Indu_Loan"] = rcon_df[domestic_ci_loan].sum(axis=1)
rcon_data["Consumer_Loan"] = rcon_df[domestic_consumer_loan].sum(axis=1)
rcon_data["Non_Rep_Loan"] = rcon_df[domestic_non_rep_loan].sum(axis=1)
rcon_data["Fed_Fund_Sold"] = rcon_df["rconb987"]
rcon_data["Reverse_Repo"] = rcon_df["rconb989"]

# Bucketed RMBS exposures
for bucket, cols in global_rmbs_buckets.items():
    existing_cols = [c for c in cols if c in rcfd_df.columns]
    rcfd_data[f"rmbs_{bucket}"] = rcfd_df[existing_cols].sum(axis=1) if existing_cols else 0

for bucket, cols in domestic_rmbs_buckets.items():
    existing_cols = [c for c in cols if c in rcon_df.columns]
    rcon_data[f"rmbs_{bucket}"] = rcon_df[existing_cols].sum(axis=1) if existing_cols else 0

# Initialize non-RMBS buckets
bucket_names = ["lt1y", "1_3y", "3_5y", "5_10y", "10_15y", "15plus"]
for bucket in bucket_names:
    rcfd_data[f"treasury_{bucket}"] = 0.0
    rcfd_data[f"other_assets_{bucket}"] = 0.0
    rcfd_data[f"res_mtg_{bucket}"] = 0.0
    rcfd_data[f"other_loan_{bucket}"] = 0.0

    rcon_data[f"treasury_{bucket}"] = 0.0
    rcon_data[f"other_assets_{bucket}"] = 0.0
    rcon_data[f"res_mtg_{bucket}"] = 0.0
    rcon_data[f"other_loan_{bucket}"] = 0.0

TREASURY_WEIGHTS = {
    "lt1y": 0.20,
    "1_3y": 0.25,
    "3_5y": 0.20,
    "5_10y": 0.20,
    "10_15y": 0.10,
    "15plus": 0.05,
}
OTHER_ASSET_WEIGHTS = {
    "lt1y": 0.10,
    "1_3y": 0.15,
    "3_5y": 0.20,
    "5_10y": 0.25,
    "10_15y": 0.20,
    "15plus": 0.10,
}
RES_MTG_WEIGHTS = {
    "lt1y": 0.05,
    "1_3y": 0.10,
    "3_5y": 0.15,
    "5_10y": 0.25,
    "10_15y": 0.25,
    "15plus": 0.20,
}
OTHER_LOAN_WEIGHTS = {
    "lt1y": 0.20,
    "1_3y": 0.20,
    "3_5y": 0.20,
    "5_10y": 0.20,
    "10_15y": 0.10,
    "15plus": 0.10,
}

# Allocate stylized non-RMBS maturities
allocate_across_buckets(rcfd_data, "security_treasury", "treasury", TREASURY_WEIGHTS)
allocate_across_buckets(rcfd_data, "security_cmbs", "other_assets", OTHER_ASSET_WEIGHTS)
allocate_across_buckets(rcfd_data, "security_abs", "other_assets", OTHER_ASSET_WEIGHTS)
allocate_across_buckets(rcfd_data, "security_other", "other_assets", OTHER_ASSET_WEIGHTS)
allocate_across_buckets(rcfd_data, "Residential_Mortgage", "res_mtg", RES_MTG_WEIGHTS)

rcfd_data["other_loan_total_tmp"] = (
    rcfd_data["Commercial_Mortgage"]
    + rcfd_data["Other_Real_Estate_Mortgage"]
    + rcfd_data["Agri_Loan"]
    + rcfd_data["Comm_Indu_Loan"]
    + rcfd_data["Consumer_Loan"]
    + rcfd_data["Non_Rep_Loan"].fillna(0)
)
allocate_across_buckets(rcfd_data, "other_loan_total_tmp", "other_loan", OTHER_LOAN_WEIGHTS)

allocate_across_buckets(rcon_data, "security_treasury", "treasury", TREASURY_WEIGHTS)
allocate_across_buckets(rcon_data, "security_cmbs", "other_assets", OTHER_ASSET_WEIGHTS)
allocate_across_buckets(rcon_data, "security_abs", "other_assets", OTHER_ASSET_WEIGHTS)
allocate_across_buckets(rcon_data, "security_other", "other_assets", OTHER_ASSET_WEIGHTS)
allocate_across_buckets(rcon_data, "Residential_Mortgage", "res_mtg", RES_MTG_WEIGHTS)

rcon_data["other_loan_total_tmp"] = (
    rcon_data["Commercial_Mortgage"]
    + rcon_data["Other_Real_Estate_Mortgage"]
    + rcon_data["Agri_Loan"]
    + rcon_data["Comm_Indu_Loan"]
    + rcon_data["Consumer_Loan"]
    + rcon_data["Non_Rep_Loan"].fillna(0)
)
allocate_across_buckets(rcon_data, "other_loan_total_tmp", "other_loan", OTHER_LOAN_WEIGHTS)

rcfd_data = rcfd_data.drop(columns=["other_loan_total_tmp"], errors="ignore")
rcon_data = rcon_data.drop(columns=["other_loan_total_tmp"], errors="ignore")

# Section 5.2: Merging rcfd and rcon to create asset tables
bank_asset = pd.merge(
    rcfd_data,
    rcon_data,
    left_index=True,
    right_index=True,
    how="outer",
    suffixes=("", "_df2"),
)

asset_df2_cols = [c for c in bank_asset.columns if c.endswith("_df2")]
for col_df2 in asset_df2_cols:
    col = col_df2.replace("_df2", "")
    if col in bank_asset.columns:
        bank_asset[col] = bank_asset[col].fillna(bank_asset[col_df2])
bank_asset = bank_asset.drop(columns=asset_df2_cols)

print("\nPost-merge bucket totals:")
for prefix in ["rmbs_", "treasury_", "other_assets_", "res_mtg_", "other_loan_"]:
    cols = [c for c in bank_asset.columns if c.startswith(prefix)]
    print(prefix, bank_asset[cols].sum().sum())

# Section 6: Creating liability tables
global_liability = pd.DataFrame(index=rcon_df.index)
global_liability["Total Liability"] = rcfd_df["rcfd2948"]
global_liability["Domestic Deposit"] = rcon_df["rcon2200"]
global_liability["Insured Deposit"] = rcon_df[insured_deposit].sum(axis=1)
global_liability["Uninsured Deposit"] = (
    global_liability["Domestic Deposit"] - global_liability["Insured Deposit"]
)
global_liability["Uninsured Time Deposits"] = rcon_df["rconj474"]
global_liability["Uninsured Long-Term Time Deposits"] = rcon_df[uninsured_long].sum(axis=1)
global_liability["Uninsured Short-Term Time Deposits"] = rcon_df["rconk222"]
global_liability["Foreign Deposit"] = rcfn_df["rcfn2200"]
global_liability["Fed Fund Purchase"] = rcon_df["rconb993"]
global_liability["Repo"] = rcon_df["rconb995"]
global_liability["Other Liability"] = rcfd_df["rcfd2930"]
global_liability["Total Equity"] = rcfd_df["rcfdg105"]
global_liability["Common Stock"] = rcfd_df["rcfd3230"]
global_liability["Preferred Stock"] = rcfd_df["rcfd3838"]
global_liability["Retained Earning"] = rcfd_df["rcfd3632"]

domestic_liability = pd.DataFrame(index=rcon_df.index)
domestic_liability["Total Liability"] = rcon_df["rcon2948"]
domestic_liability["Domestic Deposit"] = rcon_df["rcon2200"]
domestic_liability["Insured Deposit"] = rcon_df[insured_deposit].sum(axis=1)
domestic_liability["Uninsured Deposit"] = (
    domestic_liability["Domestic Deposit"] - domestic_liability["Insured Deposit"]
)
domestic_liability["Uninsured Time Deposits"] = rcon_df["rconj474"]
domestic_liability["Uninsured Long-Term Time Deposits"] = rcon_df[uninsured_long].sum(axis=1)
domestic_liability["Uninsured Short-Term Time Deposits"] = rcon_df["rconk222"]
domestic_liability["Foreign Deposit"] = rcfn_df["rcfn2200"]
domestic_liability["Fed Fund Purchase"] = rcon_df["rconb993"]
domestic_liability["Repo"] = rcon_df["rconb995"]
domestic_liability["Other Liability"] = rcon_df["rcon2930"]
domestic_liability["Total Equity"] = rcon_df["rcong105"]
domestic_liability["Common Stock"] = rcon_df["rcon3230"]
domestic_liability["Preferred Stock"] = rcon_df["rcon3838"]
domestic_liability["Retained Earning"] = rcon_df["rcon3632"]

# Section 6.2: Merging global and domestic liability tables
bank_liability = pd.merge(
    global_liability,
    domestic_liability,
    left_index=True,
    right_index=True,
    how="outer",
    suffixes=("", "_df2"),
)

liab_df2_cols = [c for c in bank_liability.columns if c.endswith("_df2")]
for col_df2 in liab_df2_cols:
    col = col_df2.replace("_df2", "")
    if col in bank_liability.columns:
        bank_liability[col] = bank_liability[col].fillna(bank_liability[col_df2])
bank_liability = bank_liability.drop(columns=liab_df2_cols)

# Save bank panel
bank_panel = bank_asset.join(bank_liability, how="outer", rsuffix="_liab")
bank_panel = bank_panel.reset_index().rename(columns={"rssd9001": "rssd_id_call"})
bank_panel["report_date"] = REPORT_DATE

bank_panel_path = DATA_DIR / f"bank_panel_{REPORT_DATE}.parquet"
bank_panel.to_parquet(bank_panel_path, index=False)
print(f"Bank panel saved -> {bank_panel_path}")

# Section 7: Summary stats for assets by bank category
threshold = 1.384e6  # $1.384 billion in thousands

bank_asset["Bank Category"] = 0
bank_asset.loc[bank_asset["Total Asset"] >= threshold, "Bank Category"] = 1

gsib_df = pull_gsib_list()
if "rssd_id_call" not in gsib_df.columns:
    if "rssd_id" in gsib_df.columns:
        gsib_df = gsib_df.rename(columns={"rssd_id": "rssd_id_call"})
    else:
        raise ValueError("GSIB list must contain 'rssd_id_call' or 'rssd_id'.")

gsib_ids = set(pd.to_numeric(gsib_df["rssd_id_call"], errors="coerce").dropna().astype(int))
bank_asset.loc[bank_asset.index.isin(gsib_ids), "Bank Category"] = 2

test_df = pd.DataFrame()
test_df["Aggregate"] = (bank_asset.sum() / bank_asset["Total Asset"].sum()) * 100
test_df["Full sample(mean)"] = (
    (bank_asset.iloc[:, :-1].div(bank_asset["Total Asset"], axis=0) * 100)
    .apply(lambda x: winsorize(x, limits=[0.05, 0.05]))
    .mean()
)
test_df["Full sample(sd)"] = (
    (bank_asset.iloc[:, :-1].div(bank_asset["Total Asset"], axis=0) * 100)
    .apply(lambda x: winsorize(x, limits=[0.05, 0.05]))
    .std()
)

bank_asset_small = bank_asset[bank_asset["Bank Category"] == 0]
test_df["small(mean)"] = (
    (bank_asset_small.iloc[:, :-1].div(bank_asset_small["Total Asset"], axis=0) * 100)
    .apply(lambda x: winsorize(x, limits=[0.05, 0.05]))
    .mean()
)
test_df["small(sd)"] = (
    (bank_asset_small.iloc[:, :-1].div(bank_asset_small["Total Asset"], axis=0) * 100)
    .apply(lambda x: winsorize(x, limits=[0.05, 0.05]))
    .std()
)

bank_asset_large = bank_asset[bank_asset["Bank Category"] == 1]
test_df["large(mean)"] = (
    (bank_asset_large.iloc[:, :-1].div(bank_asset_large["Total Asset"], axis=0) * 100)
    .apply(lambda x: winsorize(x, limits=[0.05, 0.05]))
    .mean()
)
test_df["large(sd)"] = (
    (bank_asset_large.iloc[:, :-1].div(bank_asset_large["Total Asset"], axis=0) * 100)
    .apply(lambda x: winsorize(x, limits=[0.05, 0.05]))
    .std()
)

bank_asset_gsib = bank_asset[bank_asset["Bank Category"] == 2]
test_df["GSIB(mean)"] = (
    (bank_asset_gsib.iloc[:, :-1].div(bank_asset_gsib["Total Asset"], axis=0) * 100)
    .apply(lambda x: winsorize(x, limits=[0.05, 0.05]))
    .mean()
)
test_df["GSIB(sd)"] = (
    (bank_asset_gsib.iloc[:, :-1].div(bank_asset_gsib["Total Asset"], axis=0) * 100)
    .apply(lambda x: winsorize(x, limits=[0.05, 0.05]))
    .std()
)

test_df = test_df.round(1).astype(object)

test_df.loc["Total Asset", "Aggregate"] = large_num(bank_asset["Total Asset"].sum())
test_df.loc["Total Asset", "Full sample(mean)"] = large_num(bank_asset["Total Asset"].mean())
test_df.loc["Total Asset", "Full sample(sd)"] = large_num(bank_asset["Total Asset"].std())
test_df.loc["Total Asset", "small(mean)"] = large_num(
    bank_asset[bank_asset["Bank Category"] == 0]["Total Asset"].mean()
)
test_df.loc["Total Asset", "small(sd)"] = large_num(
    bank_asset[bank_asset["Bank Category"] == 0]["Total Asset"].std()
)
test_df.loc["Total Asset", "large(mean)"] = large_num(
    bank_asset[bank_asset["Bank Category"] == 1]["Total Asset"].mean()
)
test_df.loc["Total Asset", "large(sd)"] = large_num(
    bank_asset[bank_asset["Bank Category"] == 1]["Total Asset"].std()
)
test_df.loc["Total Asset", "GSIB(mean)"] = large_num(
    bank_asset[bank_asset["Bank Category"] == 2]["Total Asset"].mean()
)
test_df.loc["Total Asset", "GSIB(sd)"] = large_num(
    bank_asset[bank_asset["Bank Category"] == 2]["Total Asset"].std()
)

test_df.loc["N Banks", "Aggregate"] = len(bank_asset)
test_df.loc["N Banks", "Full sample(mean)"] = len(bank_asset)
test_df.loc["N Banks", "small(mean)"] = (bank_asset["Bank Category"] == 0).sum()
test_df.loc["N Banks", "large(mean)"] = (bank_asset["Bank Category"] == 1).sum()
test_df.loc["N Banks", "GSIB(mean)"] = (bank_asset["Bank Category"] == 2).sum()

test_df = test_df.drop(index="Bank Category", errors="ignore")
test_df = test_df.fillna("")

test_df = test_df.rename(
    index={
        "Total Asset": "Total Asset $",
        "cash": "Cash",
        "security_total": "Securities",
        "security_treasury": "Treasury",
        "security_rmbs": "RMBS",
        "security_cmbs": "CMBS",
        "security_abs": "ABS",
        "security_other": "Other Security",
        "Total_Loan": "Total Loan",
        "Real_Estate_Loan": "Real Estate Loan",
        "Residential_Mortgage": "Residential Mortgage",
        "Commercial_Mortgage": "Commercial Mortgage",
        "Other_Real_Estate_Mortgage": "Other Real Estate Loan",
        "Agri_Loan": "Agricultural Loan",
        "Comm_Indu_Loan": "Commercial & Industrial Loan",
        "Consumer_Loan": "Consumer Loan",
        "Non_Rep_Loan": "Loan to Non-Depository",
        "Fed_Fund_Sold": "Fed Funds Sold",
        "Reverse_Repo": "Reverse Repo",
    }
)

row_order = [
    "Total Asset $",
    "N Banks",
    "Cash",
    "Securities",
    "Treasury",
    "RMBS",
    "CMBS",
    "ABS",
    "Other Security",
    "Total Loan",
    "Real Estate Loan",
    "Residential Mortgage",
    "Commercial Mortgage",
    "Other Real Estate Loan",
    "Agricultural Loan",
    "Commercial & Industrial Loan",
    "Consumer Loan",
    "Loan to Non-Depository",
    "Fed Funds Sold",
    "Reverse Repo",
]
test_df = test_df.reindex(row_order).fillna("")

# Section 8: Summary stats for liabilities by bank category
bank_liability = bank_liability.join(bank_asset[["Bank Category"]], how="left")
bank_liability = bank_liability.join(bank_asset[["Total Asset"]], how="left")
bank_liability = bank_liability.loc[:, ~bank_liability.columns.duplicated()]

df2 = pd.DataFrame()
value_cols = [c for c in bank_liability.columns if c not in ["Bank Category", "Total Asset"]]
df2["Aggregate"] = (
    bank_liability[value_cols].sum() / bank_liability["Total Asset"].sum()
) * 100

full_mean, full_sd = winsorized_mean_sd(bank_liability)
df2["Full sample(mean)"] = full_mean
df2["Full sample(sd)"] = full_sd

bank_liability_small = bank_liability[bank_liability["Bank Category"] == 0]
small_mean, small_sd = winsorized_mean_sd(bank_liability_small)
df2["small(mean)"] = small_mean
df2["small(sd)"] = small_sd

bank_liability_large = bank_liability[bank_liability["Bank Category"] == 1]
large_mean, large_sd = winsorized_mean_sd(bank_liability_large)
df2["large(mean)"] = large_mean
df2["large(sd)"] = large_sd

bank_liability_gsib = bank_liability[bank_liability["Bank Category"] == 2]
gsib_mean, gsib_sd = winsorized_mean_sd(bank_liability_gsib)
df2["GSIB(mean)"] = gsib_mean
df2["GSIB(sd)"] = gsib_sd

df2 = df2.drop(index=["Total Asset", "Bank Category"], errors="ignore")
df2 = df2.fillna(0).round(1)

# Section 9: Save summary stats to Excel
wb = Workbook()

ws_assets = wb.active
ws_assets.title = "Panel A - Assets"
write_summary_sheet(
    ws_assets, test_df, "Panel A: Asset Summary Statistics (% of Total Assets)"
)
ws_assets.column_dimensions["A"].width = 32
for col_idx in range(2, 10):
    ws_assets.column_dimensions[get_column_letter(col_idx)].width = 14

ws_liab = wb.create_sheet("Panel B - Liabilities")
write_summary_sheet(
    ws_liab, df2, "Panel B: Liability Summary Statistics (% of Total Assets)"
)
ws_liab.column_dimensions["A"].width = 36
for col_idx in range(2, 10):
    ws_liab.column_dimensions[get_column_letter(col_idx)].width = 14

summary_path = OUTPUT_DIR / f"summary_stats_{REPORT_DATE}.xlsx"
wb.save(summary_path)
print(f"Summary stats saved -> {summary_path}")

# Section 10: Save figures — Assets and Liabilities
aggregate_sum_assets = bank_asset["Total Asset"].sum()
aggregate_cash = bank_asset["cash"].sum()
aggregate_security = bank_asset["security_total"].sum()
aggregate_real_estate_loan = bank_asset["Real_Estate_Loan"].sum()
aggregate_comm_indu_loan = bank_asset["Comm_Indu_Loan"].sum()
aggregate_consumer_loan = bank_asset["Consumer_Loan"].sum()
aggregate_agri_loan = bank_asset["Agri_Loan"].sum()
aggregate_other_loan = (
    aggregate_comm_indu_loan + aggregate_consumer_loan + aggregate_agri_loan
)
aggregate_other_asset = (
    aggregate_sum_assets
    - aggregate_cash
    - aggregate_security
    - aggregate_real_estate_loan
    - aggregate_other_loan
)

proportions = {
    "Cash": (aggregate_cash / aggregate_sum_assets) * 24,
    "Security": (aggregate_security / aggregate_sum_assets) * 24,
    "Real Estate Loan": (aggregate_real_estate_loan / aggregate_sum_assets) * 24,
    "Other Loan": (aggregate_other_loan / aggregate_sum_assets) * 24,
    "Other Asset": (aggregate_other_asset / aggregate_sum_assets) * 24,
}

aggregate_liability = bank_liability["Total Liability"].sum()
aggregate_insured_deposit = bank_liability["Insured Deposit"].sum()
aggregate_uninsured_deposit = bank_liability["Uninsured Deposit"].sum()
aggregate_equity = bank_liability["Total Equity"].sum()
aggregate_other_liability = (
    aggregate_liability
    - aggregate_insured_deposit
    - aggregate_uninsured_deposit
    - aggregate_equity
)

proportions_liability = {
    "Insured Deposit": (aggregate_insured_deposit / aggregate_liability) * 24,
    "Uninsured Deposit": (aggregate_uninsured_deposit / aggregate_liability) * 24,
    "Total Equity": (aggregate_equity / aggregate_liability) * 24,
    "Other": (aggregate_other_liability / aggregate_liability) * 24,
}

colors_assets = ["navy", "royalblue", "deepskyblue", "lightskyblue", "lightblue"]
colors_liabilities = ["darkred", "red", "sandybrown", "peachpuff"]

fig, axes = plt.subplots(2, 1, figsize=(14, 5))
fig.suptitle(
    "Figure A1: U.S. Banking System Balance Sheet Composition",
    fontsize=13,
    fontweight="bold",
    y=1.01,
)

# Panel A: Assets
ax = axes[0]
starting = 0
for (category, value), color in zip(proportions.items(), colors_assets):
    ax.barh("Total Assets", value, left=starting, color=color)
    ax.text(
        starting + value / 2,
        0,
        category,
        ha="center",
        va="center",
        color="white",
        fontsize=9,
        fontweight="bold",
    )
    starting += value
ax.set_xlim(0, 24)
ax.set_xlabel("Trillion ($)", fontsize=9)
ax.set_title("Panel A: Total Assets", fontsize=10, fontweight="bold", loc="left")
ax.get_yaxis().set_visible(False)
for spine in ax.spines.values():
    spine.set_visible(False)
ax.tick_params(bottom=False, left=False)
ax.xaxis.grid(True, color="grey", linestyle="--", linewidth=0.5)

# Panel B: Liabilities
ax = axes[1]
starting_liability = 0
for (category, value), color in zip(proportions_liability.items(), colors_liabilities):
    ax.barh("Total Liability", value, left=starting_liability, color=color)
    ax.text(
        starting_liability + value / 2,
        0,
        category,
        ha="center",
        va="center",
        color="black",
        fontsize=9,
        fontweight="bold",
    )
    starting_liability += value
ax.set_xlim(0, 24)
ax.set_xlabel("Trillion ($)", fontsize=9)
ax.set_title("Panel B: Total Liabilities", fontsize=10, fontweight="bold", loc="left")
ax.get_yaxis().set_visible(False)
for spine in ax.spines.values():
    spine.set_visible(False)
ax.tick_params(bottom=False, left=False)
ax.xaxis.grid(True, color="grey", linestyle="--", linewidth=0.5)

plt.tight_layout()

figures_path = OUTPUT_DIR / f"figure_A1_{REPORT_DATE}.png"
plt.savefig(figures_path, dpi=150, bbox_inches="tight")
plt.show()
print(f"Figure saved -> {figures_path}")



# Section 11: Export summary stats to LaTeX ────────────────────────────────────

def df_to_latex(df, path, column_format=None):
    """Save a DataFrame as a booktabs-style .tex file."""
    n_cols = len(df.columns) + 1
    if column_format is None:
        column_format = "l" + "r" * (n_cols - 1)
    latex = df.to_latex(
        index=True,
        escape=True,
        column_format=column_format,
    )
    Path(path).write_text(latex)
    print(f"Saved: {path}")

    
key_rows_a = [
    "Total Asset $", "N Banks", "Cash", "Securities", "Treasury",
    "RMBS", "CMBS", "ABS", "Other Security", "Total Loan",
    "Real Estate Loan", "Residential Mortgage", "Fed Funds Sold", "Reverse Repo",
]

key_rows_b = [
    "Total Liability", "Domestic Deposit", "Insured Deposit",
    "Uninsured Deposit", "Foreign Deposit", "Fed Fund Purchase",
    "Repo", "Total Equity",
]

for sheet, key_rows, filename in [
    ("Panel A - Assets", key_rows_a, "summary_assets.tex"),
    ("Panel B - Liabilities", key_rows_b, "summary_liabilities.tex"),
]:
    df = pd.read_excel(summary_path, sheet_name=sheet, index_col=0, header=[0, 1, 2])

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [
            " ".join(str(c).strip() for c in col if "Unnamed" not in str(c)).strip()
            for col in df.columns
        ]

    df = df.reindex([r for r in key_rows if r in df.index])
    df = df.fillna("")

    df_to_latex(
        df,
        OUTPUT_DIR / filename,
        column_format="l" + "r" * len(df.columns),
    )

print("LaTeX tables exported.")
